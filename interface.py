import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from email.message import EmailMessage
import smtplib
import ssl
import threading
from sendEmail import send_email
import json
import os
from pathlib import Path


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title("Disparador de Email")
        self.geometry("900x650")
        self.minsize(850, 600)

        self.xlsx_path = ""

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        header = ctk.CTkFrame(self, corner_radius=16)
        header.grid(row=0, column=0, padx=18, pady=(18, 10), sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(header, text="Envio de Emails", font=ctk.CTkFont(size=22, weight="bold"))
        title.grid(row=0, column=0, padx=16, pady=(14, 6), sticky="w")

        subtitle = ctk.CTkLabel(
            header,
            text="Preencha credenciais, assunto, mensagem e selecione um XLSX",
            font=ctk.CTkFont(size=13),
            text_color="#A1A1AA",
        )
        subtitle.grid(row=1, column=0, padx=16, pady=(0, 14), sticky="w")

        form = ctk.CTkFrame(self, corner_radius=16)
        form.grid(row=1, column=0, padx=18, pady=10, sticky="ew")
        form.grid_columnconfigure(1, weight=1)
        form.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(form, text="Email", font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, padx=(16, 10), pady=(16, 8), sticky="w"
        )
        self.email_entry = ctk.CTkEntry(form, placeholder_text="seu_email@dominio.com")
        self.email_entry.grid(row=0, column=1, padx=(0, 16), pady=(16, 8), sticky="ew")

        ctk.CTkLabel(form, text="Senha", font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=2, padx=(16, 10), pady=(16, 8), sticky="w"
        )
        self.pass_entry = ctk.CTkEntry(form, placeholder_text="sua_senha", show="•")
        self.pass_entry.grid(row=0, column=3, padx=(0, 16), pady=(16, 8), sticky="ew")

        ctk.CTkLabel(form, text="Assunto", font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=1, column=0, padx=(16, 10), pady=8, sticky="w"
        )
        self.subject_entry = ctk.CTkEntry(form, placeholder_text="Subject do email")
        self.subject_entry.grid(row=1, column=1, columnspan=3, padx=(0, 16), pady=8, sticky="ew")

        filebox = ctk.CTkFrame(self, corner_radius=16)
        filebox.grid(row=2, column=0, padx=18, pady=10, sticky="ew")
        filebox.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(filebox, text="XLSX", font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, padx=(16, 10), pady=16, sticky="w"
        )
        self.xlsx_label = ctk.CTkLabel(filebox, text="Nenhum arquivo selecionado", text_color="#A1A1AA")
        self.xlsx_label.grid(row=0, column=1, padx=(0, 10), pady=16, sticky="w")

        self.pick_btn = ctk.CTkButton(filebox, text="Selecionar", command=self.pick_xlsx, width=140)
        self.pick_btn.grid(row=0, column=2, padx=(10, 16), pady=16, sticky="e")

        content_box = ctk.CTkFrame(self, corner_radius=16)
        content_box.grid(row=3, column=0, padx=18, pady=10, sticky="nsew")
        content_box.grid_rowconfigure(1, weight=1)
        content_box.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(content_box, text="Mensagem", font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, padx=16, pady=(16, 8), sticky="w"
        )
        self.content_text = ctk.CTkTextbox(content_box, height=220, corner_radius=12)
        self.content_text.grid(row=1, column=0, padx=16, pady=(0, 16), sticky="nsew")
        self.content_text.insert("1.0", "Olá Doutores(as)!\n\n")

        actions = ctk.CTkFrame(self, corner_radius=16)
        actions.grid(row=4, column=0, padx=18, pady=(10, 18), sticky="ew")
        actions.grid_columnconfigure(0, weight=1)

        self.status = ctk.CTkLabel(actions, text="", text_color="#A1A1AA")
        self.status.grid(row=0, column=0, padx=16, pady=16, sticky="w")

        self.send_btn = ctk.CTkButton(actions, text="Enviar", command=self.on_send, width=160)
        self.send_btn.grid(row=0, column=1, padx=16, pady=16, sticky="e")

        self.config_path = Path("config.json")
        self.load_config()


    def load_config(self):
        if not self.config_path.exists():
            return
        try:
            data = json.loads(self.config_path.read_text(encoding="utf-8"))

            self.email_entry.delete(0, "end")
            self.email_entry.insert(0, data.get("email", ""))

            self.pass_entry.delete(0, "end")
            self.pass_entry.insert(0, data.get("password", ""))

            self.subject_entry.delete(0, "end")
            self.subject_entry.insert(0, data.get("subject", ""))

            self.content_text.delete("1.0", "end")
            self.content_text.insert("1.0", data.get("content", ""))

        except Exception:
            pass

    def save_config(self, email, password, subject, content):
        try:
            data = {
                "email": email,
                "password": password,
                "subject": subject,
                "content": content,
            }
            self.config_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    def ask_save_credentials(self):
        return messagebox.askyesno("Salvar dados", "Quer salvar email, senha, assunto e mensagem neste computador?")


    def pick_xlsx(self):
        path = filedialog.askopenfilename(
            title="Selecione um XLSX",
            filetypes=[("Excel", "*.xlsx")],
        )
        if path:
            self.xlsx_path = path
            self.xlsx_label.configure(text=path)

    def on_send(self):
        email = self.email_entry.get().strip()
        password = self.pass_entry.get()
        subject = self.subject_entry.get().strip()
        content = self.content_text.get("1.0", "end").strip()

        if not email or not password or not subject or not content:
            messagebox.showerror("Erro", "Preencha email, senha, assunto e mensagem")
            return

        if not self.xlsx_path:
            messagebox.showerror("Erro", "Selecione um XLSX")
            return

        self.send_btn.configure(state="disabled")
        self.pick_btn.configure(state="disabled")
        self.status.configure(text="Enviando")

        if self.ask_save_credentials():
            self.save_config(email, password, subject, content)


        th = threading.Thread(
            target=self.worker_send,
            args=(email, password, subject, content, self.xlsx_path),
            daemon=True,
        )
        th.start()

    def worker_send(self, email, password, subject, content, xlsx_path):
        ok = 0
        fail = 0
        errors = []

        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active

            for i, row in enumerate(ws.iter_rows(values_only=True), start=2):
                to_email = row[0] if len(row) > 0 else None
                movimento = row[1] if len(row) > 1 else ""
                numero_processo = row[2] if len(row) > 2 else ""

                if to_email is None or str(to_email).strip() == "":
                    continue

                to_email = str(to_email).strip()

                # substitui variáveis no corpo
                body = content
                body = body.replace("{{ movimento }}", "" if movimento is None else str(movimento).strip())
                body = body.replace("{{ numero_processo }}", "" if numero_processo is None else str(numero_processo).strip())

                try:
                    send_email(email, password, to_email, subject, body)
                    ok += 1
                except Exception as e:
                    fail += 1
                    errors.append(f"Linha {i}: {to_email}: {type(e).__name__}: {e}")

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Erro", f"Falha ao ler XLSX: {type(e).__name__}: {e}"))
            self.after(0, self.reset_ui)
            return

        def finish():
            self.reset_ui()
            self.status.configure(text=f"Concluido. Sucesso {ok - 1}. Falhas {fail}.")
            if errors:
                messagebox.showwarning("Falhas", "\n".join(errors[:20]))

        self.after(0, finish)

    def reset_ui(self):
        self.send_btn.configure(state="normal")
        self.pick_btn.configure(state="normal")